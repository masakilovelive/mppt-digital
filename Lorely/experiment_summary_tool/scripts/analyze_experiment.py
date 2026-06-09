from __future__ import annotations

import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config" / "experiment_config.yaml"


@dataclass
class AnalysisResult:
    file_name: str
    experiment_id: str
    rows: int
    duration_s: float | None
    mode: str
    metrics: dict[str, Any]
    judgement: str
    reasons: list[str]
    graph_paths: list[Path]
    failsafe_rows: pd.DataFrame


def load_config() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Config not found: {CONFIG_PATH}")
    with CONFIG_PATH.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def rel_path(config: dict[str, Any], key: str) -> Path:
    return ROOT / config["paths"][key]


def numeric_series(df: pd.DataFrame, column: str | None) -> pd.Series | None:
    if not column or column not in df.columns:
        return None
    return pd.to_numeric(df[column], errors="coerce")


def first_text(df: pd.DataFrame, column: str | None, default: str = "") -> str:
    if not column or column not in df.columns or df.empty:
        return default
    values = df[column].dropna().astype(str)
    return values.iloc[0] if not values.empty else default


def safe_max(series: pd.Series | None) -> float | None:
    if series is None or series.dropna().empty:
        return None
    return float(series.max())


def safe_mean(series: pd.Series | None) -> float | None:
    if series is None or series.dropna().empty:
        return None
    return float(series.mean())


def safe_min(series: pd.Series | None) -> float | None:
    if series is None or series.dropna().empty:
        return None
    return float(series.min())


def rate_per_second(time_s: pd.Series | None, value: pd.Series | None) -> float | None:
    if time_s is None or value is None:
        return None
    data = pd.DataFrame({"time_s": time_s, "value": value}).dropna()
    if len(data) < 2:
        return None
    dt = data["time_s"].diff()
    dv = data["value"].diff()
    rates = (dv / dt).replace([math.inf, -math.inf], pd.NA).dropna()
    if rates.empty:
        return None
    return float(rates.max())


def time_to_reach_target(
    time_s: pd.Series | None, power: pd.Series | None, target: float, tolerance_ratio: float
) -> float | None:
    if time_s is None or power is None:
        return None
    threshold = target * (1.0 - tolerance_ratio)
    data = pd.DataFrame({"time_s": time_s, "power": power}).dropna()
    reached = data[data["power"] >= threshold]
    if reached.empty:
        return None
    start = float(data["time_s"].iloc[0])
    return float(reached["time_s"].iloc[0] - start)


def count_comm_losses(df: pd.DataFrame, column: str | None, loss_values: list[str]) -> int:
    if not column or column not in df.columns:
        return 0
    normalized = df[column].fillna("").astype(str).str.upper()
    loss_set = {v.upper() for v in loss_values}
    return int(normalized.isin(loss_set).sum())


def fault_mask(df: pd.DataFrame, column: str | None) -> pd.Series:
    if not column or column not in df.columns:
        return pd.Series(False, index=df.index)
    raw = df[column]
    numeric = pd.to_numeric(raw, errors="coerce")
    text = raw.fillna("").astype(str).str.strip().str.upper()
    return ((numeric.notna()) & (numeric != 0)) | (
        text.ne("") & ~text.isin({"0", "OK", "NONE", "NAN"})
    )


def target_power(df: pd.DataFrame, config: dict[str, Any]) -> float:
    power_target_col = config["columns"].get("power_target")
    from_csv = safe_mean(numeric_series(df, power_target_col))
    if from_csv is not None:
        return from_csv
    return float(config["thresholds"]["P_need_default"])


def analyze_csv(csv_path: Path, config: dict[str, Any]) -> AnalysisResult:
    df = pd.read_csv(csv_path)
    columns = config["columns"]
    thresholds = config["thresholds"]

    time_ms = numeric_series(df, columns.get("time"))
    time_s = time_ms / 1000.0 if time_ms is not None else None
    p_pv = numeric_series(df, columns.get("power_pv"))
    p_load = numeric_series(df, columns.get("power_load"))
    v_load = numeric_series(df, columns.get("voltage_load"))
    aux_current = numeric_series(df, columns.get("aux_current"))
    aux_energy = numeric_series(df, columns.get("aux_energy"))
    command = numeric_series(df, columns.get("command"))
    target = target_power(df, config)

    temp_series = [
        numeric_series(df, col)
        for col in columns.get("temperatures", [])
        if col in df.columns
    ]
    temp_max_values = [safe_max(s) for s in temp_series]
    temp_max_values = [v for v in temp_max_values if v is not None]
    temp_max = max(temp_max_values) if temp_max_values else None
    temp_rate_values = [rate_per_second(time_s, s) for s in temp_series]
    temp_rate_values = [v for v in temp_rate_values if v is not None]
    temp_rate_max = max(temp_rate_values) if temp_rate_values else None

    fault_rows = df[fault_mask(df, columns.get("fault"))].copy()
    comm_losses = count_comm_losses(
        df, columns.get("comm_state"), thresholds.get("comm_loss_values", [])
    )

    max_power = safe_max(p_load) or safe_max(p_pv)
    avg_power = safe_mean(p_load) or safe_mean(p_pv)
    max_current = safe_max(aux_current)
    u_cmd_rate_up = rate_per_second(time_s, command)
    p_need_time_s = time_to_reach_target(
        time_s,
        p_load if p_load is not None else p_pv,
        target,
        float(thresholds["stable_power_error_ratio"]),
    )
    overshoot = None
    if max_power is not None:
        overshoot = max(0.0, max_power - target)
    e_aux_min = safe_min(aux_energy)

    duration_s = None
    if time_s is not None and not time_s.dropna().empty:
        duration_s = float(time_s.max() - time_s.min())

    metrics = {
        "P_need": target,
        "P_max": max_power,
        "P_avg": avg_power,
        "V_load_min": safe_min(v_load),
        "I_aux_max": max_current,
        "Temp_max": temp_max,
        "Temp_rate_max_per_s": temp_rate_max,
        "u_cmd_rate_up_per_s": u_cmd_rate_up,
        "P_need_reach_time_s": p_need_time_s,
        "Overshoot_W": overshoot,
        "E_aux_min": e_aux_min,
        "Communication_loss_count": comm_losses,
        "Failsafe_count": len(fault_rows),
    }

    judgement, reasons = judge(metrics, thresholds)
    graph_paths = create_graphs(df, csv_path.stem, config)

    return AnalysisResult(
        file_name=csv_path.name,
        experiment_id=csv_path.stem,
        rows=len(df),
        duration_s=duration_s,
        mode=first_text(df, columns.get("mode"), default=""),
        metrics=metrics,
        judgement=judgement,
        reasons=reasons,
        graph_paths=graph_paths,
        failsafe_rows=fault_rows,
    )


def judge(metrics: dict[str, Any], thresholds: dict[str, Any]) -> tuple[str, list[str]]:
    fail_reasons: list[str] = []
    warn_reasons: list[str] = []

    if exceeds(metrics["P_max"], thresholds["P_max_peak"]):
        fail_reasons.append("P_max exceeded P_max_peak")
    if exceeds(metrics["I_aux_max"], thresholds["I_aux_max"]):
        fail_reasons.append("I_aux_max exceeded limit")
    if metrics["Communication_loss_count"] > 0:
        fail_reasons.append("communication loss detected")
    if metrics["Failsafe_count"] > 0:
        fail_reasons.append("failsafe/fault detected")

    if exceeds(metrics["P_avg"], thresholds["P_max_cont"]):
        warn_reasons.append("P_avg exceeded continuous limit")
    if exceeds(metrics["Temp_max"], thresholds["Temp_limit"]):
        warn_reasons.append("temperature exceeded limit")
    if metrics["V_load_min"] is not None and metrics["V_load_min"] < float(thresholds["V_load_min"]):
        warn_reasons.append("load voltage dropped below limit")
    if metrics["P_need_reach_time_s"] is None:
        warn_reasons.append("target power was not reached")
    if metrics["Overshoot_W"] is not None:
        warn_overshoot = metrics["P_need"] * float(thresholds["overshoot_warn_ratio"])
        if metrics["Overshoot_W"] > warn_overshoot:
            warn_reasons.append("power overshoot is large")

    if fail_reasons:
        return "FAIL", fail_reasons
    if warn_reasons:
        return "WARN", warn_reasons
    return "PASS", ["stable target-power response"]


def exceeds(value: Any, limit: Any) -> bool:
    return value is not None and float(value) > float(limit)


def create_graphs(df: pd.DataFrame, stem: str, config: dict[str, Any]) -> list[Path]:
    graph_dir = rel_path(config, "graph_dir")
    graph_dir.mkdir(parents=True, exist_ok=True)
    columns = config["columns"]
    time_ms = numeric_series(df, columns.get("time"))
    x = time_ms / 1000.0 if time_ms is not None else pd.Series(range(len(df)))
    xlabel = "time [s]" if time_ms is not None else "sample"

    graph_specs = [
        (
            "power",
            "Power",
            [columns.get("power_pv"), columns.get("power_load"), columns.get("power_target")],
            "W",
        ),
        ("temperature", "Temperature", columns.get("temperatures", []), "degC"),
        ("command", "Command", [columns.get("command"), columns.get("throttle")], "ratio"),
    ]

    paths: list[Path] = []
    for suffix, title, graph_columns, ylabel in graph_specs:
        existing = [col for col in graph_columns if col and col in df.columns]
        if not existing:
            continue
        fig, ax = plt.subplots(figsize=(8.0, 4.0), dpi=130)
        for col in existing:
            ax.plot(x, pd.to_numeric(df[col], errors="coerce"), label=col, linewidth=1.8)
        ax.set_title(f"{stem} - {title}")
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True, alpha=0.3)
        ax.legend(loc="best")
        fig.tight_layout()
        path = graph_dir / f"{stem}_{suffix}.png"
        fig.savefig(path)
        plt.close(fig)
        paths.append(path)
    return paths


def write_excel(results: list[AnalysisResult], config: dict[str, Any]) -> Path:
    output_dir = rel_path(config, "output_dir")
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "summary.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "01_Experiment_List"
    write_table(
        ws,
        ["experiment_id", "file", "mode", "rows", "duration_s", "judgement", "reasons"],
        [
            [
                r.experiment_id,
                r.file_name,
                r.mode,
                r.rows,
                r.duration_s,
                r.judgement,
                "; ".join(r.reasons),
            ]
            for r in results
        ],
    )

    ws = wb.create_sheet("02_Conditions")
    write_key_values(ws, flatten_dict("thresholds", config.get("thresholds", {})))

    ws = wb.create_sheet("03_Result_Summary")
    metric_names = sorted({key for r in results for key in r.metrics})
    write_table(
        ws,
        ["experiment_id", *metric_names],
        [[r.experiment_id, *[r.metrics.get(m) for m in metric_names]] for r in results],
    )

    ws = wb.create_sheet("04_Judgement")
    write_table(
        ws,
        ["experiment_id", "judgement", "reason"],
        [[r.experiment_id, r.judgement, reason] for r in results for reason in r.reasons],
    )

    ws = wb.create_sheet("05_Graphs")
    row = 1
    for r in results:
        ws.cell(row=row, column=1, value=r.experiment_id)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for graph_path in r.graph_paths:
            if graph_path.exists():
                img = ExcelImage(str(graph_path))
                img.width = 640
                img.height = 320
                ws.add_image(img, f"A{row}")
                row += 18

    ws = wb.create_sheet("06_Failsafe_History")
    rows: list[list[Any]] = []
    headers: list[str] = ["experiment_id"]
    for r in results:
        if r.failsafe_rows.empty:
            continue
        for _, row_data in r.failsafe_rows.iterrows():
            rows.append([r.experiment_id, *row_data.tolist()])
        headers = ["experiment_id", *r.failsafe_rows.columns.tolist()]
    write_table(ws, headers, rows)

    ws = wb.create_sheet("07_Final_Parameters")
    write_key_values(ws, config.get("final_parameters", {}))

    ws = wb.create_sheet("08_Notes")
    write_table(
        ws,
        ["item", "note"],
        [
            ["raw_data", "Keep original CSV files unchanged in logs/raw_csv."],
            ["analysis", "Re-run scripts/analyze_experiment.py after adding CSV logs."],
            ["review", "Edit judgement thresholds in config/experiment_config.yaml."],
        ],
    )

    for sheet in wb.worksheets:
        style_sheet(sheet)

    wb.save(path)
    return path


def write_table(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def write_key_values(ws: Any, data: dict[str, Any]) -> None:
    write_table(ws, ["key", "value"], [[key, value] for key, value in data.items()])


def flatten_dict(prefix: str, data: dict[str, Any]) -> dict[str, Any]:
    flat: dict[str, Any] = {}
    for key, value in data.items():
        joined = f"{prefix}.{key}"
        if isinstance(value, dict):
            flat.update(flatten_dict(joined, value))
        elif isinstance(value, list):
            flat[joined] = ", ".join(str(v) for v in value)
        else:
            flat[joined] = value
    return flat


def style_sheet(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)
    ws.freeze_panes = "A2"


def write_markdown_report(results: list[AnalysisResult], config: dict[str, Any]) -> Path:
    output_dir = rel_path(config, "output_dir")
    path = output_dir / "report.md"
    lines = [
        "# Experiment Report",
        "",
        "| Experiment | Judgement | P_need | P_max | P_avg | Temp_max | Reasons |",
        "| --- | --- | ---: | ---: | ---: | ---: | --- |",
    ]
    for r in results:
        m = r.metrics
        lines.append(
            "| {exp} | {judgement} | {p_need} | {p_max} | {p_avg} | {temp} | {reasons} |".format(
                exp=r.experiment_id,
                judgement=r.judgement,
                p_need=fmt(m.get("P_need")),
                p_max=fmt(m.get("P_max")),
                p_avg=fmt(m.get("P_avg")),
                temp=fmt(m.get("Temp_max")),
                reasons="; ".join(r.reasons),
            )
        )
    lines.append("")
    lines.append("## Graphs")
    lines.append("")
    for r in results:
        lines.append(f"### {r.experiment_id}")
        for graph in r.graph_paths:
            rel = graph.relative_to(output_dir).as_posix()
            lines.append(f"![{graph.stem}]({rel})")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def write_pdf_report(markdown_path: Path, config: dict[str, Any]) -> Path | None:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
    except ImportError:
        return None

    output_dir = rel_path(config, "output_dir")
    pdf_path = output_dir / "report.pdf"
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    for line in markdown_path.read_text(encoding="utf-8").splitlines():
        if line.startswith("# "):
            story.append(Paragraph(line[2:], styles["Title"]))
            story.append(Spacer(1, 12))
        elif line.startswith("## "):
            story.append(Paragraph(line[3:], styles["Heading2"]))
        elif line.startswith("|") or line.startswith("![") or not line.strip():
            continue
        else:
            story.append(Paragraph(line, styles["BodyText"]))
    doc.build(story)
    return pdf_path


def fmt(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float):
        return f"{value:.3g}"
    return str(value)


def main() -> int:
    config = load_config()
    input_dir = rel_path(config, "input_dir")
    output_dir = rel_path(config, "output_dir")
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_files = sorted(input_dir.glob("*.csv"))
    if not csv_files:
        print(f"No CSV files found in {input_dir}")
        return 1

    results = [analyze_csv(csv_path, config) for csv_path in csv_files]
    excel_path = write_excel(results, config)
    markdown_path = write_markdown_report(results, config)
    pdf_path = write_pdf_report(markdown_path, config)

    print(f"Analyzed {len(results)} CSV file(s)")
    print(f"Excel: {excel_path}")
    print(f"Markdown: {markdown_path}")
    if pdf_path:
        print(f"PDF: {pdf_path}")
    else:
        print("PDF: skipped because reportlab is not installed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
